import pygame
import openpyxl
import curses
from pygame import mixer
from openpyxl import load_workbook

column_a_values = []

def main(stdscr):
    # Initialize Pygame
    pygame.init()

    # Initialize Pygame mixer
    pygame.mixer.init()
    # Clear screen
    stdscr.clear()

    # Load the workbook and get the sheet names
    wb = openpyxl.load_workbook('test.xlsx')
    sheet_names = wb.sheetnames

    current_index = 0

    while True:
        # Display the sheet names
        stdscr.clear()
        for idx, name in enumerate(sheet_names):
            if idx == current_index:
                stdscr.addstr(idx, 0, name, curses.A_REVERSE)
            else:
                stdscr.addstr(idx, 0, name)

        key = stdscr.getch()

        # Navigate through the sheet names
        if key == curses.KEY_UP and current_index > 0:
            current_index -= 1
        elif key == curses.KEY_DOWN and current_index < len(sheet_names) - 1:
            current_index += 1
        elif key == curses.KEY_ENTER or key in [10, 13]:
            break

    selected_sheet = sheet_names[current_index]
    worksheet = wb[selected_sheet]
    max_row = worksheet.max_row
    delay_s = worksheet['G1'].value * 1000
    stdscr.clear()
    stdscr.addstr(0, 20, f"NOW PLAYING {selected_sheet}")
    stdscr.addstr(1, 20, f"DELAY {delay_s} SECONDS")
    stdscr.refresh()
    # Iterate over all cells in column A and load/play their values

    # Collect all values in column A
    for row in range(1, max_row + 1):
        cell_value = worksheet[f'A{row}'].value
        if cell_value is not None:
            column_a_values.append(cell_value)
    # Play each music file sequentially
    for file_path in column_a_values:
        #play_music(file_path)
        pygame.mixer.music.load(file_path)
        pygame.mixer.music.play()
        pygame.time.wait(delay_s)  # 3 seconds (3000 milliseconds)
        # Wait until music finishes playing
        while pygame.mixer.music.get_busy():
            pygame.time.Clock().tick(10)  # Adjust the tick rate as needed

    # Wait a bit before quitting
    pygame.time.wait(1000)  # 3 seconds (3000 milliseconds)
    pygame.quit()
    stdscr.clear()

    # Get the value of cell A1 from the selected sheet
    stdscr.addstr(1, 0, f"PLAY FINISH")

    stdscr.refresh()
    stdscr.getch()

curses.wrapper(main)








